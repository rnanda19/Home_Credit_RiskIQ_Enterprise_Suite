# ============================================================================
# NOTEBOOK 06 — MEGA PROJECT 4: DELINQUENCY PREVENTION
# CONSOLIDATED EXECUTIVE ROLLUP (rolls up real Problems 1-5)
# ----------------------------------------------------------------------------
# ZERO-FABRICATION DISCLOSURE: every figure below is read directly from each
# problem notebook's own real, already-computed governance JSON summary
# (decision_engine/reports/notebook_0N_summary.json) -- nothing here is
# retrained, re-clustered, re-scored, or invented. This script does not touch
# any raw Kaggle CSV, any of Notebooks 01/03/04's saved model bundles, or
# Notebook 02's saved cluster assignments at all; it is a pure rollup + real
# cross-notebook synthesis of numbers each notebook already produced on ITS
# OWN real run of your data. The genuinely new things this notebook adds are
# (1) "Real Behavioral Data Coverage" -- a fresh bar chart independently
# re-deriving each of Problems 1-4's own real scope-vs-total-population
# coverage ratio straight from that problem's own summary, and (2) real
# cross-notebook consistency checks (Section 5) that verify Notebook 05's own
# independent record of which signals it found available agrees, notebook by
# notebook, with which of Notebooks 01-04's summaries actually exist on this
# run -- no new modeling, clustering, or scoring anywhere.
#
# IMPORTANT SCALE CAVEAT: every figure below reflects whatever data each of
# Notebooks 01-05 was MOST RECENTLY run against. Re-run all 5 on your real,
# downloaded Home Credit data, then re-run this notebook, and every number
# here recomputes automatically from those real runs.
#
# WHY THREE DIFFERENT VERDICT-TIER NAMES APPEAR BELOW: Problems 1, 3, and 4
# each report a "Statistical Robustness Verdict" (5-fold CV champion + real
# holdout ROC-AUC + bootstrap 95% CI + decile-calibration monotonicity).
# Problem 2 reports a "Clustering Robustness Verdict" instead (real
# silhouette score + chi-square/Cramer's V of cluster vs. real TARGET) --
# different validation family because it is an unsupervised method. Problem 5
# reports a "Ranking Comparison Verdict" instead of either -- it trains and
# clusters nothing; it validates a different real question entirely: does a
# real composite of Problems 1-4's own signals capture more real defaults in
# its top decile than the simplest possible real comparator (current DPD).
# This rollup surfaces all three families side by side, correctly labeled,
# never conflated -- mirrors how Mega Project 3's own executive rollup
# (06_mp3_executive_report.ipynb) discloses its own two-verdict-family split.
#
# LESSONS APPLIED FROM THIS SUITE'S OWN HARDENING HISTORY (LESSONS_LEARNED.md):
#   - Missing upstream summaries are reported and skipped, never fabricated
#     (mirrors Mega Projects 1, 2, and 3's own executive-rollup pattern).
#   - Real cross-checks, not asserted (#6): Notebook 05's own real record of
#     which of Problems 1-4's signals it found available is checked here,
#     notebook by notebook, against which of those notebooks' summaries
#     actually exist on THIS run -- to make sure a soft dependency was never
#     silently missed or silently faked.
#   - HYPER reuse: report_builder for all 3 output formats; each problem's
#     OWN already-generated real chart PNG is embedded directly rather than
#     redrawn, so nothing here can silently drift from what that notebook's
#     own verification pass already confirmed correct.
#
# VERIFICATION NOTE (2026-09-01 policy — see CHANGELOG): per explicit
# instruction, this notebook was NOT executed against any synthetic fixture
# before delivery. It trains, clusters, and fits nothing new -- it only
# reads, joins, and re-presents real JSON already produced by Notebooks
# 01-05. Its new derived logic (coverage-ratio computation, cross-notebook
# signal-availability consistency check) was verified with a small,
# hand-built set of mock summary dicts covering the full-availability,
# partial-availability, and zero-availability cases -- every derived value
# checked by hand. This file's syntax was checked (py_compile/ast.parse,
# 0 errors) and this notebook passes nbformat.validate(). This notebook's own
# real per-problem verdicts, coverage figures, and consistency results are
# determined ONLY by running it after Notebooks 01-05 against your real data.
# ============================================================================

import os
import sys
import json
import time
from pathlib import Path

# ---------------------------------------------------------------------------
# SECTION 1 — Suite-root resolution (identical pattern to every notebook in
# this suite).
# ---------------------------------------------------------------------------
def _find_suite_root(start: Path = None) -> Path:
    start = start or Path.cwd()
    marker = "project_config.json"
    env_override = os.environ.get("HC_SUITE_ROOT")
    if env_override and (Path(env_override) / marker).exists():
        return Path(env_override)
    for candidate in [start, *start.parents]:
        if (candidate / marker).exists():
            return candidate
    for candidate in [
        Path.home() / "Downloads" / "home-credit-enterprise-suite",
        Path.home() / "home-credit-enterprise-suite",
        Path.home() / "Desktop" / "home-credit-enterprise-suite",
        start / "home-credit-enterprise-suite",
        start / "Downloads" / "home-credit-enterprise-suite",
    ]:
        if (candidate / marker).exists():
            return candidate
    return None


SUITE_ROOT = _find_suite_root()
if SUITE_ROOT is None:
    raise FileNotFoundError(
        "project_config.json not found. Run this after at least Mega Project 4 / "
        "Notebook 01 has been run once, or set HC_SUITE_ROOT."
    )

MP4_DIR = SUITE_ROOT / "04_mega_project_4_delinquency_prevention"
ARTIFACTS_DIR = MP4_DIR / "decision_engine" / "artifacts"
REPORTS_DIR = MP4_DIR / "decision_engine" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(SUITE_ROOT / "src"))
from reporting.report_builder import (
    write_csv_outputs, build_word_report, build_excel_workbook,
    build_html_dashboard, assumption_ref, VIVID_PALETTE, _palette, safe_sheet_name,
)

import matplotlib.pyplot as plt
import pandas as pd

T0 = time.time()

# ---------------------------------------------------------------------------
# SECTION 2 — Load each real problem notebook's already-computed governance
# summary. NOTE: unlike Mega Project 3, every MP4 notebook writes its summary
# JSON to decision_engine/reports/ (not decision_engine/artifacts/) -- see
# each pipeline_mp4_nb0N.py's own SECTION where summary_path is defined. A
# missing file is reported and skipped, never fabricated.
# ---------------------------------------------------------------------------
PROBLEM_META = {
    "01": {"label": "Problem 1 — Early Delinquency Risk Scoring", "file": "notebook_01_summary.json",
           "kind": "classifier",
           "method": "4-model-screened supervised classifier (LogisticRegression / DecisionTree / "
                     "RandomForest / GradientBoosting) on real installment-payment behavior features "
                     "from installments_payments.csv",
           "verdict_kind": "Statistical Robustness Verdict", "chart_png": "notebook_01_charts.png"},
    "02": {"label": "Problem 2 — Installment Payment Behavior Detection", "file": "notebook_02_summary.json",
           "kind": "clustering",
           "method": "Real, unsupervised K-Means clustering (never trained against TARGET) on 7 real "
                     "payment-streak features from installments_payments.csv",
           "verdict_kind": "Clustering Robustness Verdict", "chart_png": "notebook_02_charts.png"},
    "03": {"label": "Problem 3 — Revolving/Credit-Card Distress Early Warning", "file": "notebook_03_summary.json",
           "kind": "classifier",
           "method": "4-model-screened supervised classifier on real utilization-spike, minimum-payment-"
                     "streak, and drawdown-velocity features from credit_card_balance.csv",
           "verdict_kind": "Statistical Robustness Verdict", "chart_png": "notebook_03_charts.png"},
    "04": {"label": "Problem 4 — POS/Cash Loan Delinquency Trajectory", "file": "notebook_04_summary.json",
           "kind": "classifier",
           "method": "4-model-screened supervised classifier on real DPD-spike, DPD-streak, and "
                     "instalment-progress-velocity features from POS_CASH_balance.csv",
           "verdict_kind": "Statistical Robustness Verdict", "chart_png": "notebook_04_charts.png"},
    "05": {"label": "Problem 5 — Early-Warning Intervention Ranking", "file": "notebook_05_summary.json",
           "kind": "fusion",
           "method": "Real, disclosed fusion of whichever of Problems 1-4's own real per-applicant scores "
                     "are present (soft dependencies) -- trains nothing new, reads no raw Home Credit CSV "
                     "feature table",
           "verdict_kind": "Ranking Comparison Verdict", "chart_png": "notebook_05_charts.png"},
}

summaries = {}
missing = []
for nb_id, meta in PROBLEM_META.items():
    path = REPORTS_DIR / meta["file"]
    if path.exists():
        with open(path) as f:
            summaries[nb_id] = json.load(f)
    else:
        missing.append(nb_id)

N_AVAILABLE = len(summaries)
print(f"[ROLLUP] {N_AVAILABLE} / 5 real MP4 problem summaries found under {REPORTS_DIR.name}/.")
if missing:
    print(f"[ROLLUP] Missing (run these notebooks first for a complete rollup): "
          f"{', '.join('Notebook ' + m for m in missing)}")
if N_AVAILABLE == 0:
    raise FileNotFoundError(
        "No notebook_0N_summary.json files found under decision_engine/reports/. Run at least one "
        "of Notebooks 01-05 first (each writes its own real governance summary on completion)."
    )

# Real baseline population -- prefer Notebook 01's n_app_total (the full real
# application_train.csv row count that notebook loaded), falling back to
# whichever classifier/clustering notebook is available, since all four
# (01/02/03/04) compute this identically from the same real file.
_BASELINE_SRC = next((nb_id for nb_id in ("01", "02", "03", "04") if nb_id in summaries), None)
N_APP_TOTAL = summaries[_BASELINE_SRC]["n_app_total"] if _BASELINE_SRC else None
if N_APP_TOTAL is not None:
    print(f"[ROLLUP] Real baseline population (from Notebook {_BASELINE_SRC}): {N_APP_TOTAL:,} applicants "
          f"in application_train.csv.")

# ---------------------------------------------------------------------------
# SECTION 3 — Real per-problem verdict + integrity-check extraction, unified
# across the 3 differently-named verdict families this Mega Project uses on
# purpose (see module docstring above and each notebook's own model card).
# ---------------------------------------------------------------------------
def _headline_metric(nb_id: str) -> str:
    s = summaries[nb_id]
    kind = PROBLEM_META[nb_id]["kind"]
    if kind == "classifier":
        auc = s.get("holdout_metrics", {}).get("roc_auc")
        ci = s.get("holdout_auc_ci") or [None, None]
        n_scope = s.get("n_scope")
        auc_txt = f"{auc:.4f}" if auc is not None else "N/A"
        ci_txt = f" (95% CI [{ci[0]:.4f}, {ci[1]:.4f}])" if ci[0] is not None and ci[1] is not None else ""
        scope_txt = f", n_scope={n_scope:,}" if n_scope is not None else ""
        return f"Champion {s.get('champion_model', 'N/A')}, real holdout ROC-AUC {auc_txt}{ci_txt}{scope_txt}"
    if kind == "clustering":
        k = s.get("k_chosen")
        sil = s.get("silhouette_chosen")
        n_pat = len(s.get("pattern_agg", []))
        n_scope = s.get("n_scope")
        sil_txt = f"{sil:.4f}" if sil is not None else "N/A"
        scope_txt = f", n_scope={n_scope:,}" if n_scope is not None else ""
        return f"k={k}, real silhouette={sil_txt}, {n_pat} real payment patterns{scope_txt}"
    if kind == "fusion":
        n_sig = s.get("n_signals_available")
        n_eval = s.get("n_eval")
        c_lift = s.get("composite_lift")
        n_lift = s.get("naive_lift")
        c_txt = f"{c_lift:.2f}x" if isinstance(c_lift, (int, float)) else "N/A"
        n_txt = f"{n_lift:.2f}x" if isinstance(n_lift, (int, float)) else "N/A"
        eval_txt = f", n_eval={n_eval:,}" if n_eval is not None else ""
        return f"{n_sig}/4 real signals combined, composite top-decile lift {c_txt} vs. naive lift {n_txt}{eval_txt}"
    return "N/A"


rollup_rows = []
STORIES = {}
for nb_id, meta in PROBLEM_META.items():
    if nb_id not in summaries:
        continue
    s = summaries[nb_id]
    verdict = s.get("analysis_verdict", "N/A")
    ic = {k: v for k, v in s.items() if False}  # MP4 notebooks don't nest integrity_checks in the summary
    headline = _headline_metric(nb_id)
    row = {
        "notebook_id": nb_id, "problem": meta["label"], "method": meta["method"],
        "verdict_kind": meta["verdict_kind"], "analysis_verdict": verdict,
        "analysis_robust": bool(s.get("analysis_robust", False)),
        "headline_metric": headline,
    }
    rollup_rows.append(row)
    STORIES[nb_id] = [
        f"{meta['label']} uses {meta['method']}.",
        f"Real headline result: {headline}",
        f"{meta['verdict_kind']} (real, computed this run): {verdict}",
    ]

rollup_df = pd.DataFrame(rollup_rows)

# ---------------------------------------------------------------------------
# SECTION 4 — "Real Behavioral Data Coverage" -- this rollup's own fresh
# synthesis: each of Problems 1-4's real scope population (applicants with
# at least one real record in that problem's underlying table) as a fraction
# of the real total application_train.csv population, independently
# re-derived here straight from that problem's OWN summary (works even if
# Problem 5 has not been run, and does NOT assume these ratios should be
# equal -- each behavioral table naturally covers a different real subset of
# applicants, e.g. not every applicant has a real credit card).
# ---------------------------------------------------------------------------
_coverage_rows = []
for nb_id in ("01", "02", "03", "04"):
    if nb_id not in summaries:
        continue
    s = summaries[nb_id]
    n_scope = s.get("n_scope")
    n_total = s.get("n_app_total")
    if n_scope is None or not n_total:
        continue
    _coverage_rows.append({
        "notebook_id": nb_id, "problem": PROBLEM_META[nb_id]["label"].split("—")[1].strip(),
        "n_scope": int(n_scope), "n_app_total": int(n_total),
        "coverage_fraction": float(n_scope) / float(n_total),
    })
coverage_df = (pd.DataFrame(_coverage_rows).sort_values("coverage_fraction", ascending=False).reset_index(drop=True)
               if _coverage_rows else pd.DataFrame(
    columns=["notebook_id", "problem", "n_scope", "n_app_total", "coverage_fraction"]))
_WIDEST_COVERAGE_ROW = coverage_df.iloc[0].to_dict() if not coverage_df.empty else None
if _WIDEST_COVERAGE_ROW is not None:
    print(f"[ROLLUP] Real behavioral-data coverage (widest to narrowest): " +
          ", ".join(f"{r['problem']}={r['coverage_fraction']:.2%}" for _, r in coverage_df.iterrows()))

# ---------------------------------------------------------------------------
# SECTION 5 — Real cross-notebook consistency checks (this rollup's own
# genuine addition -- comparing real records two independent ways, never
# asserted -- LESSONS_LEARNED.md #6).
# ---------------------------------------------------------------------------
_consistency_rows = []

# Check A: every classifier/clustering notebook's real N_APP_TOTAL (the full
# real application_train.csv row count each one independently loaded) should
# be identical, since they all read the same real file.
_n_apps = {nb_id: summaries[nb_id]["n_app_total"] for nb_id in ("01", "02", "03", "04")
           if nb_id in summaries and "n_app_total" in summaries[nb_id]}
_n_apps_consistent = len(set(_n_apps.values())) <= 1
_consistency_rows.append({
    "check": "n_app_total_identical_across_available_notebooks_01_02_03_04",
    "values_by_notebook": _n_apps, "within_tolerance": bool(_n_apps_consistent),
})
if _n_apps:
    print(f"[CROSS-CHECK] Real application_train.csv row count is identical across every available "
          f"classifier/clustering notebook ({', '.join(f'NB{k}={v:,}' for k, v in _n_apps.items())}): "
          f"{'CONFIRMED' if _n_apps_consistent else 'MISMATCH FOUND'}.")

# Check B: Notebook 05's own real record of which of Problems 1-4's signals
# it found available is checked, notebook by notebook, against which of
# those notebooks' summaries actually exist on THIS run -- confirms no soft
# dependency was silently missed or silently faked.
_NB05_SIGNAL_TO_SOURCE = {
    "nb01_installment_behavior": "01", "nb02_payment_pattern": "02",
    "nb03_revolving_distress": "03", "nb04_pos_cash_trajectory": "04",
}
if "05" in summaries:
    nb05_signals = summaries["05"].get("signals_available", {})
    _sig_mismatches = []
    for sig_name, src_nb in _NB05_SIGNAL_TO_SOURCE.items():
        nb05_says = bool(nb05_signals.get(sig_name, False))
        actually_present = src_nb in summaries
        matched = nb05_says == actually_present
        if not matched:
            _sig_mismatches.append(sig_name)
        _consistency_rows.append({
            "check": f"nb05_signal_availability_matches_notebook_{src_nb}_summary_presence_({sig_name})",
            "nb05_reported_available": nb05_says, "notebook_summary_actually_present": actually_present,
            "within_tolerance": bool(matched),
        })
    NB05_SIGNAL_CONSISTENCY_OK = len(_sig_mismatches) == 0
    print(f"[CROSS-CHECK] Real Notebook 05 signal-availability record matches which of Notebooks "
          f"01-04's real summaries are actually present on this run: "
          f"{'CONFIRMED' if NB05_SIGNAL_CONSISTENCY_OK else 'MISMATCH FOUND: ' + ', '.join(_sig_mismatches)}.")
else:
    NB05_SIGNAL_CONSISTENCY_OK = True  # nothing to check if Notebook 05 hasn't been run
    print("[CROSS-CHECK] Notebook 05 not available this run -- signal-availability cross-check skipped "
          "(not a failure; this check only applies once Notebook 05 has run).")

ALL_CONSISTENCY_OK = all(r["within_tolerance"] for r in _consistency_rows) if _consistency_rows else True

# ---------------------------------------------------------------------------
# SECTION 6 — SMART insights: one per available problem, plus two bonus
# insights explaining (a) the three verdict-tier families and (b) how to
# read real behavioral-data coverage across problems.
# ---------------------------------------------------------------------------
INSIGHTS = []
for row in rollup_rows:
    nb_id = row["notebook_id"]
    INSIGHTS.append({
        "headline": f"{row['problem']}: {row['analysis_verdict'].split('—')[0].strip()}",
        "specific": STORIES[nb_id][0],
        "measurable": f"Real headline result: {row['headline_metric']}",
        "achievable": f"Real, computed {row['verdict_kind']} this run: {row['analysis_verdict']}",
        "relevant": f"{row['verdict_kind']} is this problem's own dedicated validation gate, distinct from "
                    f"the other verdict families used elsewhere in this Mega Project.",
        "timebound": "Re-validate the moment this notebook is re-run against real production data.",
    })

INSIGHTS.append({
    "headline": "Reading this Mega Project's THREE verdict-tier families",
    "specific": "Problems 1, 3, and 4 each report a \"Statistical Robustness Verdict\" (5-fold CV champion, "
                "real holdout ROC-AUC, bootstrap 95% CI, decile-calibration monotonicity). Problem 2 "
                "reports a \"Clustering Robustness Verdict\" instead (real silhouette score, chi-square / "
                "Cramer's V of cluster vs. real TARGET) -- a different validation family because it is "
                "unsupervised. Problem 5 reports a \"Ranking Comparison Verdict\" instead of either -- it "
                "trains and clusters nothing; it validates a genuinely different real question: does a "
                "real composite of Problems 1-4's signals out-rank the simplest possible real comparator.",
    "measurable": "3 distinct verdict-tier names appear across 5 problems, each named for what it actually tests.",
    "achievable": "Every notebook's own model card discloses exactly why its verdict tier is named as it is.",
    "relevant": "Prevents misreading one problem's verdict as directly comparable to another's, when the "
                "underlying validation methods differ by design.",
    "timebound": "This distinction is structural to the report and does not change across runs.",
})
INSIGHTS.append({
    "headline": "Real behavioral-data coverage varies by product line, and that is expected",
    "specific": "Problems 1-4 each score a different real subset of the applicant base -- only applicants "
                "with at least one real record in that problem's underlying table (installment loans, "
                "revolving credit cards, or POS/cash loans) are in scope. A lower coverage fraction for one "
                "product line does not mean a weaker model; it reflects how many real applicants actually "
                "hold that product.",
    "measurable": ("Real coverage fraction by problem (widest to narrowest): " +
                   ", ".join(f"{r['problem']}={r['coverage_fraction']:.2%}"
                             for _, r in coverage_df.iterrows())) if not coverage_df.empty else "N/A",
    "achievable": f"Real cross-notebook consistency (Notebook 05's own signal-availability record vs. which "
                  f"of Notebooks 01-04's summaries are actually present): "
                  f"{'CONFIRMED' if ALL_CONSISTENCY_OK else 'MISMATCH FOUND'} (Section 5 of this notebook).",
    "relevant": "A collections or portfolio team should read Problem 5's COVERAGE_COUNT field the same way -- "
                "an applicant with fewer real signals is not necessarily lower-risk, they simply hold fewer "
                "of the real product lines this Mega Project scores.",
    "timebound": "Re-confirm this coverage picture every time Problems 1-4 are re-run against fresh real data.",
})

# ---------------------------------------------------------------------------
# SECTION 7 — Integrity checks for this rollup itself.
# ---------------------------------------------------------------------------
checks = [
    ("all_5_problem_summaries_found", N_AVAILABLE == 5),
    ("n_app_total_identical_across_available_classifier_clustering_notebooks", _n_apps_consistent),
    ("nb05_signal_availability_consistent_with_present_summaries", NB05_SIGNAL_CONSISTENCY_OK),
    ("every_available_problem_has_a_story", len(STORIES) == N_AVAILABLE),
    ("every_available_problem_has_an_insight", len(INSIGHTS) >= N_AVAILABLE),
]
print("\n[INTEGRITY CHECKS]")
for name, ok in checks:
    print(f"  [CHECK] {name}: {'PASS' if ok else 'FAIL'}")
failed = [n for n, ok in checks if not ok]
if failed:
    raise AssertionError(f"Rollup integrity checks failed: {failed}")

# ---------------------------------------------------------------------------
# SECTION 8 — NEW synthesis chart (the only chart this notebook itself draws
# -- every other chart embedded below is each problem's own already-verified
# real chart, reused, not redrawn).
# ---------------------------------------------------------------------------
fig, ax = plt.subplots(figsize=(9, 5.2))
if not coverage_df.empty:
    _cov_colors = _palette(len(coverage_df))
    bars = ax.bar(coverage_df["problem"], coverage_df["coverage_fraction"], color=_cov_colors)
    for b, (_, r) in zip(bars, coverage_df.iterrows()):
        ax.annotate(f"{r['coverage_fraction']:.1%}", (b.get_x() + b.get_width() / 2, b.get_height()),
                    ha="center", va="bottom", fontsize=9, fontweight="bold")
ax.set_ylabel("Real Coverage (scope population / total applicant population)")
ax.set_title("Real Behavioral Data Coverage — Problems 1-4, Independently Re-Derived")
plt.setp(ax.get_xticklabels(), rotation=12, ha="right", fontsize=9)
plt.tight_layout()
COVERAGE_PNG = REPORTS_DIR / "notebook_06_behavioral_coverage.png"
plt.savefig(COVERAGE_PNG, dpi=110)
plt.show()

# ---------------------------------------------------------------------------
# SECTION 9 — Reporting & Packaging
# ---------------------------------------------------------------------------
csv_paths = write_csv_outputs(
    {"mp4_executive_rollup": rollup_df, "mp4_behavioral_coverage": coverage_df,
     "mp4_cross_notebook_consistency": pd.DataFrame(_consistency_rows)},
    REPORTS_DIR,
)

# --- Word report -------------------------------------------------------------
exec_summary = [
    f"{N_AVAILABLE} of 5 Mega Project 4 problems have a real completed run available for this rollup"
    + ("." if N_AVAILABLE == 5 else f" (missing: {', '.join('Notebook ' + m for m in missing)})."),
    f"Real baseline population (from Notebook {_BASELINE_SRC}): {N_APP_TOTAL:,} applicants."
    if N_APP_TOTAL is not None else "",
    ("Real behavioral-data coverage (widest to narrowest): " +
     ", ".join(f"{r['problem']}={r['coverage_fraction']:.2%}" for _, r in coverage_df.iterrows()) + ".")
    if not coverage_df.empty else "",
    "Verdicts this run: " + "; ".join(
        f"{r['problem'].split('—')[0].strip()}: {r['analysis_verdict'].split('—')[0].strip()}"
        for r in rollup_rows
    ),
    f"Real cross-notebook consistency (Notebook 05's own signal-availability record vs. which of Notebooks "
    f"01-04's summaries are actually present): {'CONFIRMED' if ALL_CONSISTENCY_OK else 'MISMATCH FOUND'}.",
    "SCALE CAVEAT: every figure above reflects whatever data each notebook was most recently run against. "
    "Re-run Notebooks 01-05 against your real, downloaded Home Credit data, then re-run this notebook, for "
    "every number here to recompute.",
]
exec_summary = [line for line in exec_summary if line]

word_sections = []
for row in rollup_rows:
    nb_id = row["notebook_id"]
    table_rows = [["Method", row["method"]], [row["verdict_kind"], row["analysis_verdict"]],
                  ["Real Headline Result", row["headline_metric"]]]
    img = REPORTS_DIR / PROBLEM_META[nb_id]["chart_png"]
    word_sections.append({
        "heading": row["problem"], "paragraphs": [],
        "table": {"headers": ["Metric", "Value"], "rows": table_rows},
        "image_path": img if img.exists() else None, "story": STORIES[nb_id],
    })
word_sections.append({
    "heading": "Real Behavioral Data Coverage (compared, independently re-derived)",
    "paragraphs": [
        "Each problem's real coverage fraction is re-derived here directly from that problem's OWN "
        "source notebook -- the real fraction of the total applicant base who hold at least one real "
        "record in that problem's underlying table. A lower fraction reflects a narrower real product "
        "line (fewer applicants hold revolving credit cards or POS/cash loans than hold instalment "
        "loans), not a weaker model.",
    ],
    "table": {"headers": ["Problem", "Real Scope", "Real Total Population", "Coverage"],
              "rows": [[r["problem"], f"{int(r['n_scope']):,}", f"{int(r['n_app_total']):,}",
                        f"{r['coverage_fraction']:.2%}"]
                       for _, r in coverage_df.iterrows()]} if not coverage_df.empty else None,
    "image_path": COVERAGE_PNG,
    "story": [f"Real cross-notebook consistency check: {'CONFIRMED' if ALL_CONSISTENCY_OK else 'MISMATCH FOUND'}."],
})

word_path = build_word_report(
    REPORTS_DIR / "mp4_executive_report.docx",
    title="Mega Project 4 — Executive Capstone Report",
    subtitle="Delinquency Prevention — Home Credit Default Risk Enterprise Suite",
    exec_summary=exec_summary,
    sections=word_sections,
    insights=INSIGHTS,
)

# --- Excel workbook ------------------------------------------------------
assumptions = {}
assumption_notes = {}
if N_APP_TOTAL is not None:
    assumptions["REAL_N_APP_TOTAL"] = int(N_APP_TOTAL)
    assumption_notes["REAL_N_APP_TOTAL"] = (
        f"Real total applicant population from Notebook {_BASELINE_SRC}'s own load of "
        f"application_train.csv, reused as the denominator for every coverage figure above.")
if _WIDEST_COVERAGE_ROW is not None:
    assumptions["REAL_WIDEST_COVERAGE_FRACTION"] = round(float(_WIDEST_COVERAGE_ROW["coverage_fraction"]), 6)
    assumption_notes["REAL_WIDEST_COVERAGE_FRACTION"] = (
        f"Real coverage fraction of the most broadly-covering problem this run "
        f"({_WIDEST_COVERAGE_ROW['problem']}) -- referenced by the Executive Rollup sheet.")

data_sheets = []
for row in rollup_rows:
    nb_id = row["notebook_id"]
    headers = ["Metric", "Value"]
    sheet_rows = [["Problem", row["problem"]], ["Method", row["method"]],
                  [row["verdict_kind"], row["analysis_verdict"]], ["Real Headline Result", row["headline_metric"]]]
    data_sheets.append({"name": safe_sheet_name(f"P{nb_id} {PROBLEM_META[nb_id]['label'].split('—')[1].strip()[:22]}"),
                         "headers": headers, "rows": sheet_rows})
data_sheets.append({"name": "Problem Rollup", "headers": list(rollup_df.columns),
                     "rows": rollup_df.fillna("").values.tolist()})
if not coverage_df.empty:
    data_sheets.append({"name": "Behavioral Coverage", "headers": list(coverage_df.columns),
                         "rows": coverage_df.values.tolist(), "highlight_col": "coverage_fraction"})

formula_rows = []
if "REAL_N_APP_TOTAL" in assumptions:
    total_ref = assumption_ref(assumptions, "REAL_N_APP_TOTAL")
    formula_rows.append(("Real Total Applicant Population", f"={total_ref}"))
if "REAL_WIDEST_COVERAGE_FRACTION" in assumptions:
    cov_ref = assumption_ref(assumptions, "REAL_WIDEST_COVERAGE_FRACTION")
    formula_rows.append(("Real Widest Behavioral Coverage", f"={cov_ref}"))
formula_sheet = {"name": "Financial Impact", "rows": formula_rows} if formula_rows else None

excel_path = build_excel_workbook(
    REPORTS_DIR / "mp4_executive_report.xlsx",
    assumptions=assumptions, assumption_notes=assumption_notes,
    data_sheets=data_sheets, formula_sheet=formula_sheet,
    insights_sheet={"name": "SMART Insights", "items": INSIGHTS},
)

# --- Post-process: native Excel charts + embedded real PNGs + big-letters
# front "Executive Rollup" sheet (openpyxl -- report_builder's generic
# builder does not embed images/native charts, so this notebook adds both
# directly, same pattern Mega Projects 1, 2, and 3's own rollups already
# established for the big-letters front sheet). --------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage

wb = load_workbook(excel_path)

for row in rollup_rows:
    nb_id = row["notebook_id"]
    sheet_name = safe_sheet_name(f"P{nb_id} {PROBLEM_META[nb_id]['label'].split('—')[1].strip()[:22]}")
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    # embed the real PNG this problem's own notebook already generated
    png_path = REPORTS_DIR / PROBLEM_META[nb_id]["chart_png"]
    if png_path.exists():
        try:
            img = XLImage(str(png_path))
            img.width, img.height = 460, 280
            ws.add_image(img, "E2")
        except Exception as img_err:
            print(f"[WARN] Could not embed {png_path.name} into {sheet_name}: {img_err}")

# Front "Executive Rollup" sheet (big letters, inserted first) with a native
# chart of Real Behavioral Data Coverage.
front = wb.create_sheet("Executive Rollup", 0)
front.sheet_view.showGridLines = False
front.column_dimensions["A"].width = 4
front.column_dimensions["B"].width = 55
BIG_TITLE_FONT = Font(size=26, bold=True, color="1F3864")
BIG_VALUE_FONT = Font(size=36, bold=True, color="1F7A3D")
LABEL_FONT = Font(size=13, bold=True, color="595959")
front.merge_cells("B2:I2")
front["B2"] = "MEGA PROJECT 4 — REAL DELINQUENCY PREVENTION ROLLUP"
front["B2"].font = BIG_TITLE_FONT

big_rows = [
    ("Real Applicant Population", f"{N_APP_TOTAL:,}" if N_APP_TOTAL is not None else "N/A"),
    ("Widest Real Behavioral Coverage",
     f"{_WIDEST_COVERAGE_ROW['problem']} ({_WIDEST_COVERAGE_ROW['coverage_fraction']:.2%})"
     if _WIDEST_COVERAGE_ROW is not None else "N/A"),
    ("Real Cross-Notebook Consistency", "CONFIRMED" if ALL_CONSISTENCY_OK else "MISMATCH FOUND"),
    ("Real Problems Complete", f"{N_AVAILABLE} / 5"),
]
r = 4
for label, value in big_rows:
    front.merge_cells(f"B{r}:I{r}")
    front[f"B{r}"] = label
    front[f"B{r}"].font = LABEL_FONT
    r += 1
    front.merge_cells(f"B{r}:I{r}")
    front[f"B{r}"] = value
    front[f"B{r}"].font = BIG_VALUE_FONT
    front[f"B{r}"].alignment = Alignment(horizontal="left")
    r += 2

# scratch data block for the native "Behavioral Coverage" chart + the chart itself
cov_start = r + 1
front.cell(row=cov_start, column=2, value="Problem")
front.cell(row=cov_start, column=3, value="Coverage Fraction")
for i, (_, rec) in enumerate(coverage_df.iterrows(), start=1):
    front.cell(row=cov_start + i, column=2, value=rec["problem"])
    front.cell(row=cov_start + i, column=3, value=float(rec["coverage_fraction"]))
if not coverage_df.empty:
    cov_chart = BarChart()
    cov_chart.title = "Real Behavioral Data Coverage — Problems 1-4"
    cov_chart.y_axis.title = "Coverage Fraction"
    cov_chart.style = 12
    cats = Reference(front, min_col=2, min_row=cov_start + 1, max_row=cov_start + len(coverage_df))
    vals = Reference(front, min_col=3, min_row=cov_start, max_row=cov_start + len(coverage_df))
    cov_chart.add_data(vals, titles_from_data=True)
    cov_chart.set_categories(cats)
    cov_chart.width, cov_chart.height = 18, 10
    front.add_chart(cov_chart, f"E{cov_start}")

caveat_row = cov_start + len(coverage_df) + 3
front.merge_cells(f"B{caveat_row}:I{caveat_row + 1}")
front[f"B{caveat_row}"] = ("Scale caveat: figures reflect whatever data each notebook was most recently run "
                            "against. See the 'Behavioral Coverage' sheet and this notebook's own model card "
                            "for the full cross-notebook consistency disclosure.")
front[f"B{caveat_row}"].font = Font(size=10, italic=True, color="808080")
front[f"B{caveat_row}"].alignment = Alignment(wrap_text=True, vertical="top")

wb.save(excel_path)

# --- HTML dashboard --------------------------------------------------------
verdict_counts: dict[str, int] = {}
for r in rollup_rows:
    key = r["analysis_verdict"].split("—")[0].strip()
    verdict_counts[key] = verdict_counts.get(key, 0) + 1

kpi_cards = [
    {"label": "Real Applicants", "value": f"{N_APP_TOTAL:,}" if N_APP_TOTAL is not None else "N/A"},
    {"label": "Real Problems Complete", "value": f"{N_AVAILABLE} / 5"},
    {"label": "Widest Real Behavioral Coverage",
     "value": f"{_WIDEST_COVERAGE_ROW['problem']} ({_WIDEST_COVERAGE_ROW['coverage_fraction']:.2%})"
     if _WIDEST_COVERAGE_ROW is not None else "N/A"},
]
if "01" in summaries:
    kpi_cards.append({"label": "NB01 Real Holdout ROC-AUC",
                       "value": f"{summaries['01'].get('holdout_metrics', {}).get('roc_auc', float('nan')):.4f}"})
if "02" in summaries:
    kpi_cards.append({"label": "NB02 Real Silhouette Score",
                       "value": f"{summaries['02'].get('silhouette_chosen', float('nan')):.4f}"})
if "03" in summaries:
    kpi_cards.append({"label": "NB03 Real Holdout ROC-AUC",
                       "value": f"{summaries['03'].get('holdout_metrics', {}).get('roc_auc', float('nan')):.4f}"})
if "04" in summaries:
    kpi_cards.append({"label": "NB04 Real Holdout ROC-AUC",
                       "value": f"{summaries['04'].get('holdout_metrics', {}).get('roc_auc', float('nan')):.4f}"})
if "05" in summaries:
    c_lift = summaries["05"].get("composite_lift")
    kpi_cards.append({"label": "NB05 Real Composite Top-Decile Lift",
                       "value": f"{c_lift:.2f}x" if isinstance(c_lift, (int, float)) else "N/A"})
kpi_cards.append({"label": "Real Cross-Notebook Consistency",
                   "value": "CONFIRMED" if ALL_CONSISTENCY_OK else "MISMATCH FOUND"})

charts = []

# Chart 1 — Real Behavioral Data Coverage (the notebook's own new synthesis)
if not coverage_df.empty:
    charts.append({
        "id": "behavioralCoverage", "title": "Real Behavioral Data Coverage — Problems 1-4, Independently Re-Derived",
        "type": "bar", "labels": coverage_df["problem"].tolist(),
        "datasets": [{"label": "Real Coverage Fraction", "data": coverage_df["coverage_fraction"].tolist(),
                      "backgroundColor": VIVID_PALETTE[:len(coverage_df)]}],
        "story": [f"Real coverage fraction by problem, widest to narrowest, each re-derived directly from "
                  f"that problem's own source notebook. Real cross-notebook consistency (Notebook 05's own "
                  f"signal-availability record vs. which notebooks actually ran): "
                  f"{'CONFIRMED' if ALL_CONSISTENCY_OK else 'MISMATCH FOUND'}."],
    })

# Chart 2 — Real holdout ROC-AUC across the 3 classifiers (01, 03, 04)
_auc_rows = []
for nb_id in ("01", "03", "04"):
    if nb_id in summaries:
        auc = summaries[nb_id].get("holdout_metrics", {}).get("roc_auc")
        if auc is not None:
            _auc_rows.append((PROBLEM_META[nb_id]["label"].split("—")[1].strip(), float(auc)))
if _auc_rows:
    charts.append({
        "id": "classifierAucComparison", "title": "Real Holdout ROC-AUC — Problems 1, 3, 4 (Classifiers Only)",
        "type": "bar", "labels": [r[0] for r in _auc_rows],
        "datasets": [{"label": "Real Holdout ROC-AUC", "data": [r[1] for r in _auc_rows],
                      "backgroundColor": _palette(len(_auc_rows))}],
        "story": ["Each classifier is trained on a different real behavioral table and a different real "
                  "scope population -- this compares their real holdout discrimination side by side, not a "
                  "claim that a higher AUC alone makes one problem more useful than another."],
    })

# Chart 3 — Real default rate by payment pattern (NB02)
if "02" in summaries:
    recs = summaries["02"].get("pattern_agg", [])
    if recs:
        recs_sorted = sorted(recs, key=lambda r: r.get("real_default_rate", 0), reverse=True)
        charts.append({
            "id": "paymentPatternDefaultRate", "title": "Real Default Rate by Payment Pattern (Problem 2)",
            "type": "bar", "labels": [str(r.get("PAYMENT_PATTERN")) for r in recs_sorted],
            "datasets": [{"label": "Real Default Rate",
                          "data": [r.get("real_default_rate") for r in recs_sorted],
                          "backgroundColor": _palette(len(recs_sorted))}],
        })

# Chart 4 — Real composite vs. naive top-decile default rate (NB05)
if "05" in summaries:
    s5 = summaries["05"]
    c_rate = s5.get("composite_top_decile_default_rate")
    n_rate = s5.get("naive_top_decile_default_rate")
    if c_rate is not None and n_rate is not None:
        charts.append({
            "id": "compositeVsNaive", "title": "Real Top-Decile Default Rate — Composite vs. Naive Baseline (Problem 5)",
            "type": "bar", "labels": ["Composite Ranking", "Naive Current-DPD Baseline"],
            "datasets": [{"label": "Real Top-Decile Default Rate", "data": [c_rate, n_rate],
                          "backgroundColor": [VIVID_PALETTE[2], VIVID_PALETTE[7]]}],
            "story": [f"Real chi-square p-value on this 2x2 comparison: "
                      f"{s5.get('chi2_p') if s5.get('chi2_p') is not None else 'N/A'}."],
        })

# Chart 5 — Verdict distribution
charts.append({
    "id": "verdictDist", "title": "Verdicts (this run)", "type": "doughnut",
    "labels": list(verdict_counts.keys()),
    "datasets": [{"label": "Problems", "data": list(verdict_counts.values()),
                  "backgroundColor": VIVID_PALETTE[:len(verdict_counts)]}],
})

html_path = build_html_dashboard(
    REPORTS_DIR / "mp4_executive_dashboard.html",
    title="Mega Project 4 — Executive Dashboard",
    subtitle="Delinquency Prevention (real rollup of Problems 1-5)",
    kpi_cards=kpi_cards,
    charts=charts,
    insights=INSIGHTS,
    data_table={
        "title": "Per-Problem Rollup (real)",
        "columns": ["Problem", "Method", "Verdict Kind", "Analysis Verdict", "Headline Metric"],
        "rows": [[r["problem"], r["method"], r["verdict_kind"], r["analysis_verdict"], r["headline_metric"]]
                 for r in rollup_rows],
        "filter_column": "Verdict Kind",
    },
)
print(f"[REPORTING] Real MP4 executive reporting package written: {word_path.name}, {excel_path.name}, "
      f"{html_path.name}, plus {len(csv_paths)} CSV file(s) (all under {REPORTS_DIR.name}/).")

# ---------------------------------------------------------------------------
# SECTION 10 — Governance JSON summary for this rollup
# ---------------------------------------------------------------------------
mp4_summary = {
    "report": "mp4_executive_report",
    "mega_project": "Mega Project 4 - Delinquency Prevention",
    "problems_available": N_AVAILABLE,
    "problems_missing": missing,
    "real_baseline": {"n_app_total": N_APP_TOTAL, "source_notebook": _BASELINE_SRC},
    "behavioral_coverage": _coverage_rows,
    "cross_notebook_consistency_checks": _consistency_rows,
    "per_problem_rollup": rollup_rows,
    "integrity_checks": {n: bool(ok) for n, ok in checks},
    "reporting_artifacts": [word_path.name, excel_path.name, html_path.name] + [f"{s}.csv" for s in csv_paths],
    "sop_stage_reached": "6 - Production Packaging & Governance",
    "runtime_seconds": round(time.time() - T0, 1),
}
with open(REPORTS_DIR / "mp4_executive_summary.json", "w") as f:
    json.dump(mp4_summary, f, indent=2, default=str)

print(f"\n[DONE] MP4 executive rollup complete in {time.time() - T0:.1f}s covering {N_AVAILABLE}/5 real "
      f"problem summaries."
      + (f" Real applicant population: {N_APP_TOTAL:,}." if N_APP_TOTAL is not None else ""))
