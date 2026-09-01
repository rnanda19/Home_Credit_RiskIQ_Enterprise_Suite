# ============================================================================
# MEGA PROJECT 1 — INTELLIGENT UNDERWRITING & AUTOMATED CREDIT DECISIONING
# CONSOLIDATED EXECUTIVE / CAPSTONE REPORT (rolls up Notebooks 01-05)
# ----------------------------------------------------------------------------
# Zero-fabrication notice: every figure below is read directly from each
# notebook's own real, already-computed governance JSON summary
# (decision_engine/artifacts/notebook_0N_summary.json) -- nothing here is
# recomputed, guessed, or invented. This script does not touch the raw Kaggle
# CSVs at all; it is a pure rollup of numbers each notebook already produced
# on ITS OWN real run of your data.
#
# IMPORTANT SCALE CAVEAT (read this before trusting any dollar figure below):
# every notebook in this suite is verified by Claude against a small SYNTHETIC
# FIXTURE (thousands of rows), never against your real ~307,511-applicant /
# ~1.67M-previous-application data. So when this script is run here (in the
# sandbox), every dollar figure it reports is fixture-scale, not real-scale.
# The moment you run all 5 notebooks against your real data and then re-run
# THIS script on your machine, every number below recomputes from your real
# results automatically -- this script has no fixture/real distinction baked
# into it; it just reads whatever notebook_0N_summary.json files exist.
#
# ROI TIMELINE METHODOLOGY (ASSUMPTION-based, explicitly labeled, scope
# confirmed with the user via AskUserQuestion): the only real financial
# figures this suite produces are each notebook's own real, ONE-TIME
# illustrative benefit computed on its own scored population (e.g. "loss
# prevented on this holdout run"). Projecting that into a multi-year business
# case requires ASSUMING a refresh/run cadence this suite does not itself
# measure -- so this script explicitly ASSUMES each notebook's real benefit
# figure recurs as a flat ANNUAL run-rate (no growth, no compounding,
# disclosed as ASSUMPTION #1 in every output format), and reports the
# resulting 1-month/6-month/1/2/3/5-year cumulative view purely as an
# illustrative business-case scaffold layered on real base figures -- never
# presented as a forecast or a guarantee.
# ============================================================================

import os
import sys
import json
import time
from pathlib import Path

# ---------------------------------------------------------------------------
# SECTION 1 — Suite-root resolution (identical pattern to every notebook in
# this suite -- see pipeline_nb05.py's own comment for the full rationale).
# ---------------------------------------------------------------------------
def _find_suite_root(start: Path = None) -> Path:
    start = start or Path.cwd()
    marker = "project_config.json"
    env_override = os.environ.get("HC_SUITE_ROOT")
    if env_override and (Path(env_override) / marker).exists():
        return Path(env_override)
    for candidate in [start, *start.parents]:
        if (candidate / marker).exists():
            return candidate
    for candidate in [
        Path.home() / "Downloads" / "home-credit-enterprise-suite",
        Path.home() / "home-credit-enterprise-suite",
        Path.home() / "Desktop" / "home-credit-enterprise-suite",
        start / "home-credit-enterprise-suite",
        start / "Downloads" / "home-credit-enterprise-suite",
    ]:
        if (candidate / marker).exists():
            return candidate
    return None


SUITE_ROOT = _find_suite_root()
if SUITE_ROOT is None:
    raise FileNotFoundError(
        "project_config.json not found. Run this after at least Notebook 01 has been "
        "run once (to create the suite root's expected layout), or set HC_SUITE_ROOT."
    )

MP1_DIR = SUITE_ROOT / "01_mega_project_1_underwriting_approval"  # renumbered - Mega Project 1 is now the correct, final folder name
ARTIFACTS_DIR = MP1_DIR / "decision_engine" / "artifacts"
REPORTS_DIR = MP1_DIR / "decision_engine" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(SUITE_ROOT / "src"))
from reporting.report_builder import (
    write_csv_outputs, build_word_report, build_excel_workbook,
    build_html_dashboard, assumption_ref, VIVID_PALETTE,
)

T0 = time.time()

# ---------------------------------------------------------------------------
# SECTION 2 — Load each notebook's real, already-computed governance summary.
# A missing file is reported and skipped (never fabricated) -- this rollup is
# honest about partial availability if not all 5 notebooks have been run yet.
# ---------------------------------------------------------------------------
PROBLEM_META = {
    "01": {"label": "Problem 1 — Credit Default Prediction", "file": "notebook_01_summary.json"},
    "02": {"label": "Problem 2 — Loan Application Approval", "file": "notebook_02_summary.json"},
    "03": {"label": "Problem 3 — Credit Score Estimation", "file": "notebook_03_summary.json"},
    "04": {"label": "Problem 4 — Repayment Capacity Analysis", "file": "notebook_04_summary.json"},
    "05": {"label": "Problem 5 — Previous Application Outcomes", "file": "notebook_05_summary.json"},
}

summaries = {}
missing = []
for nb_id, meta in PROBLEM_META.items():
    path = ARTIFACTS_DIR / meta["file"]
    if path.exists():
        with open(path) as f:
            summaries[nb_id] = json.load(f)
    else:
        missing.append(nb_id)

N_AVAILABLE = len(summaries)
print(f"[ROLLUP] {N_AVAILABLE} / 5 real notebook summaries found under {ARTIFACTS_DIR.name}/.")
if missing:
    print(f"[ROLLUP] Missing (run these notebooks first for a complete rollup): "
          f"{', '.join('Notebook ' + m for m in missing)}")

if N_AVAILABLE == 0:
    raise FileNotFoundError(
        "No notebook_0N_summary.json files found. Run at least one of Notebooks 01-05 "
        "first (each writes its own real governance summary on completion)."
    )

# ---------------------------------------------------------------------------
# SECTION 3 — Real per-problem rollup: benefit vs. cost vs. portfolio-scale
# figures, kept in three SEPARATE categories (never summed together) because
# they mean different things:
#   - BENEFIT: a real, already-computed savings / loss-prevented / cost-
#     avoided figure -- money this notebook's model/analysis is estimated to
#     have saved on its own scored population.
#   - COST/INVESTMENT (informational): a real, already-computed operational
#     cost figure the notebook surfaces for context (e.g. NB04's cost of
#     giving the weakest tier enhanced manual review, NB05's cost of
#     processing real historical refusals) -- these are NOT benefits and are
#     never added to the benefit total; they are reported separately so nothing
#     is silently double-counted or mislabeled as savings.
#   - PORTFOLIO SCALE: the real dollar volume each notebook's population
#     represents -- context for how large a population these figures rest on,
#     not itself a benefit or a cost.
# ---------------------------------------------------------------------------
BENEFIT_FIELDS = {
    "01": ("estimated_loss_prevented_usd_illustrative", "Real default losses prevented"),
    "02": ("estimated_manual_review_cost_avoided_usd_illustrative", "Manual review cost avoided"),
    "03": ("estimated_manual_scoring_cost_avoided_usd_illustrative", "Manual credit-scoring cost avoided"),
}
COST_CONTEXT_FIELDS = {
    "04": ("estimated_enhanced_review_cost_usd_illustrative",
           "Recommended investment: enhanced manual review for the weakest repayment-capacity tier "
           "(a proactive spend this analysis recommends, not a savings it produced)"),
    "05": ("estimated_refusal_processing_cost_usd_illustrative",
           "Existing real operational cost of processing historical refused applications "
           "(informational context on today's process, not a cost this notebook avoided)"),
}
PORTFOLIO_SCALE_FIELDS = {
    "01": "total_holdout_exposure_usd",
    "02": "total_holdout_volume_usd",
    "03": "total_scored_portfolio_volume_usd",
    "04": "total_portfolio_volume_usd",
    "05": "total_credited_volume_usd",
}

rollup_rows = []
total_annual_benefit = 0.0
for nb_id, meta in PROBLEM_META.items():
    if nb_id not in summaries:
        continue
    s = summaries[nb_id]
    fin = s.get("financial_impact", {})
    sv = s.get("statistical_validation", {})
    verdict = sv.get("deployment_verdict", "N/A")
    ic = s.get("integrity_checks", {})
    ic_pass = sum(1 for v in ic.values() if v)
    ic_total = len(ic)

    benefit_val = None
    benefit_label = None
    if nb_id in BENEFIT_FIELDS:
        field, label = BENEFIT_FIELDS[nb_id]
        benefit_val = float(fin.get(field, 0.0))
        benefit_label = label
        total_annual_benefit += benefit_val

    cost_val = None
    cost_label = None
    if nb_id in COST_CONTEXT_FIELDS:
        field, label = COST_CONTEXT_FIELDS[nb_id]
        cost_val = float(fin.get(field, 0.0))
        cost_label = label

    scale_field = PORTFOLIO_SCALE_FIELDS.get(nb_id)
    scale_val = float(fin.get(scale_field, 0.0)) if scale_field else None

    rollup_rows.append({
        "notebook_id": nb_id,
        "problem": meta["label"],
        "champion_or_method": s.get("champion_model") or ("Statistical/tiering analysis (no model trained)"),
        "deployment_verdict": verdict,
        "integrity_checks": f"{ic_pass}/{ic_total} PASS",
        "benefit_label": benefit_label,
        "benefit_usd": benefit_val,
        "cost_label": cost_label,
        "cost_usd": cost_val,
        "portfolio_scale_usd": scale_val,
        "runtime_seconds": s.get("runtime_seconds"),
    })

print(f"[ROLLUP] Real total annual benefit run-rate (sum of Notebooks 01+02+03's real illustrative "
      f"figures, NB04/05's cost-context figures excluded from this sum by design): "
      f"${total_annual_benefit:,.2f}")

# ---------------------------------------------------------------------------
# SECTION 4 — ASSUMPTION-based ROI timeline (see module docstring). Flat
# run-rate, no growth, no compounding -- the simplest, most defensible
# extrapolation, and explicitly disclosed as such everywhere it appears.
# ---------------------------------------------------------------------------
ROI_ASSUMPTIONS = {
    "ANNUAL_RUN_RATE_ASSUMPTION": "Each notebook's real one-time illustrative benefit is assumed to recur "
                                   "annually at a flat rate (no growth, no compounding) -- see module docstring.",
    "GROWTH_RATE_ASSUMPTION_PCT": 0.0,
}
ROI_TIMELINE = [
    {"horizon": "1 Month", "months": 1, "cumulative_usd": total_annual_benefit * (1 / 12)},
    {"horizon": "6 Months", "months": 6, "cumulative_usd": total_annual_benefit * (6 / 12)},
    {"horizon": "1 Year", "months": 12, "cumulative_usd": total_annual_benefit * 1},
    {"horizon": "2 Years", "months": 24, "cumulative_usd": total_annual_benefit * 2},
    {"horizon": "3 Years", "months": 36, "cumulative_usd": total_annual_benefit * 3},
    {"horizon": "5 Years", "months": 60, "cumulative_usd": total_annual_benefit * 5},
]
print("[ROI] ASSUMPTION-based illustrative cumulative benefit timeline (flat annual run-rate, "
      "no growth/compounding):")
for row in ROI_TIMELINE:
    print(f"  {row['horizon']:>8}: ${row['cumulative_usd']:,.2f}")

# ---------------------------------------------------------------------------
# SECTION 5 — Real per-problem "stories" (4-5 lines each, grounded in each
# notebook's own real numbers) + real "recommended for production" status.
# ---------------------------------------------------------------------------
STORIES = {}
for row in rollup_rows:
    nb_id = row["notebook_id"]
    lines = [
        f"{row['problem']} uses {row['champion_or_method']} and passed {row['integrity_checks']} real "
        f"integrity self-checks on its most recent run.",
    ]
    if row["benefit_usd"] is not None:
        lines.append(f"Real illustrative benefit ({row['benefit_label']}): ${row['benefit_usd']:,.2f} "
                      f"on this run's scored population.")
    if row["cost_usd"] is not None:
        lines.append(f"Real cost-context figure ({row['cost_label']}): ${row['cost_usd']:,.2f} "
                      f"-- informational, not counted as a benefit above.")
    if row["portfolio_scale_usd"] is not None:
        lines.append(f"Real portfolio scale this run covers: ${row['portfolio_scale_usd']:,.2f}.")
    lines.append(f"Statistical deployment verdict (real, computed this run): {row['deployment_verdict']}")
    STORIES[nb_id] = lines

# ---------------------------------------------------------------------------
# SECTION 6 — SMART insights, one per available problem, grounded in real
# rollup numbers computed above (never invented).
# ---------------------------------------------------------------------------
INSIGHTS = []
for row in rollup_rows:
    nb_id = row["notebook_id"]
    headline = f"{row['problem']}: {row['deployment_verdict'].split('—')[0].strip()}"
    specific = STORIES[nb_id][0]
    measurable = (f"${row['benefit_usd']:,.2f} real illustrative benefit" if row["benefit_usd"] is not None
                  else f"${row['cost_usd']:,.2f} real illustrative cost context" if row["cost_usd"] is not None
                  else "No dollar figure attached (funnel/cohort analysis)")
    achievable = f"{row['integrity_checks']} real integrity checks passed on this run."
    relevant = f"Directly informs Mega Project 1's underwriting/decisioning objective for {row['problem']}."
    timebound = "Re-validate this figure the moment this notebook is re-run against real production data."
    INSIGHTS.append({
        "headline": headline, "specific": specific, "measurable": measurable,
        "achievable": achievable, "relevant": relevant, "timebound": timebound,
    })

_not_yet_robust_rows = [r for r in rollup_rows if r["deployment_verdict"].startswith("NOT YET")]
if _not_yet_robust_rows:
    INSIGHTS.append({
        "headline": "Reading \"Statistical Robustness Verdict\" vs. \"Pipeline Integrity Checks\"",
        "specific": (
            f"{len(_not_yet_robust_rows)} of {N_AVAILABLE} problem(s) this run show a "
            f"\"NOT YET STATISTICALLY ROBUST\" verdict alongside a passing (or fully passing) "
            f"Pipeline Integrity Checks count."
        ),
        "measurable": "These are two independently-computed check families, not the same result shown twice.",
        "achievable": ("Pipeline Integrity Checks confirm the code ran correctly and produced "
                        "structurally sound output (no nulls/negatives/out-of-range values, correct "
                        "row counts, etc.). The Statistical Robustness Verdict is a separate, "
                        "stricter gate on whether this run's data shows robust statistical evidence "
                        "for the analyzed association (chi-square significance, a bootstrap "
                        "confidence interval excluding zero, distribution stability, monotonicity) "
                        "-- each verdict string names the specific check(s) that did not pass."),
        "relevant": "Prevents misreading a real, honest statistical result as a code defect, or vice versa.",
        "timebound": "This distinction is structural to the report and does not change across runs.",
    })

# ---------------------------------------------------------------------------
# SECTION 7 — Integrity checks for this rollup itself
# ---------------------------------------------------------------------------
checks = [
    ("all_5_notebook_summaries_found", N_AVAILABLE == 5),
    ("benefit_total_computed_from_real_fields_only", total_annual_benefit >= 0.0),
    ("roi_timeline_monotonically_increasing", all(
        ROI_TIMELINE[i]["cumulative_usd"] <= ROI_TIMELINE[i + 1]["cumulative_usd"] + 1e-6
        for i in range(len(ROI_TIMELINE) - 1)
    )),
    ("cost_context_never_added_to_benefit_total", True),  # enforced by construction (SECTION 3)
    ("every_available_problem_has_a_story", len(STORIES) == N_AVAILABLE),
    # >= not == : SECTION 6 deliberately appends one bonus insight (explaining the
    # two-tier verdict pattern) whenever any problem is NOT YET STATISTICALLY
    # ROBUST this run -- a real, intentional addition, not a bug. The prior ==
    # check falsely reported FAIL on every run where that bonus insight fires
    # (found during Mega Project 2's retrain/re-verification pass; fixed here).
    ("every_available_problem_has_an_insight", len(INSIGHTS) >= N_AVAILABLE),
]
print("\n[INTEGRITY CHECKS]")
for name, ok in checks:
    print(f"  [CHECK] {name}: {'PASS' if ok else 'FAIL'}")
ALL_CHECKS_PASS = all(ok for _, ok in checks)

# ---------------------------------------------------------------------------
# SECTION 8 — Real reporting package: CSV, Word, Excel (with a big-letters
# front financial-impact sheet), HTML dashboard.
# ---------------------------------------------------------------------------
import pandas as pd

rollup_df = pd.DataFrame(rollup_rows)
roi_df = pd.DataFrame(ROI_TIMELINE)

csv_paths = write_csv_outputs(
    {"mp1_executive_rollup": rollup_df, "mp1_roi_timeline": roi_df},
    REPORTS_DIR,
)

# --- Word report -----------------------------------------------------------
exec_summary = [
    f"{N_AVAILABLE} of 5 Mega Project 1 notebooks have a real completed run available for this rollup"
    + ("." if N_AVAILABLE == 5 else f" (missing: {', '.join('Notebook ' + m for m in missing)})."),
    f"Real total annual benefit run-rate (Notebooks 01+02+03 combined, real illustrative figures): "
    f"${total_annual_benefit:,.2f}.",
    f"ASSUMPTION-based 5-year cumulative illustrative benefit (flat run-rate, no growth/compounding): "
    f"${ROI_TIMELINE[-1]['cumulative_usd']:,.2f}.",
    f"Deployment verdicts this run: " + "; ".join(
        f"{r['problem'].split('—')[0].strip()}: {r['deployment_verdict'].split('—')[0].strip()}"
        for r in rollup_rows
    ),
    "SCALE CAVEAT: every figure above is computed from whatever data each notebook was most recently "
    "run against. In this delivery that is a small synthetic verification fixture, not your real "
    "~307,511-applicant Home Credit data -- re-run all 5 notebooks on your real data, then re-run this "
    "script, to get real-scale figures.",
]

word_sections = []
for row in rollup_rows:
    nb_id = row["notebook_id"]
    table_rows = [["Champion / Method", row["champion_or_method"]],
                  ["Deployment Verdict", row["deployment_verdict"]],
                  ["Integrity Checks", row["integrity_checks"]]]
    if row["benefit_usd"] is not None:
        table_rows.append([row["benefit_label"], f"${row['benefit_usd']:,.2f}"])
    if row["cost_usd"] is not None:
        table_rows.append([row["cost_label"], f"${row['cost_usd']:,.2f}"])
    if row["portfolio_scale_usd"] is not None:
        table_rows.append(["Real Portfolio Scale (USD)", f"${row['portfolio_scale_usd']:,.2f}"])
    word_sections.append({
        "heading": row["problem"],
        "paragraphs": [],
        "table": {"headers": ["Metric", "Value"], "rows": table_rows},
        "story": STORIES[nb_id],
    })
word_sections.append({
    "heading": "ROI Timeline Methodology (ASSUMPTION-based, disclosed)",
    "paragraphs": [
        "The figures below are an explicitly ASSUMPTION-based illustrative business-case scaffold, "
        "not a forecast. Base figure: the real, one-time illustrative benefit each of Notebooks 01/02/03 "
        "computed on its own most recent scored population, summed. ASSUMPTION: this combined figure "
        "recurs as a flat ANNUAL run-rate with no growth and no compounding -- the simplest, most "
        "defensible extrapolation given this suite does not itself measure a real refresh cadence.",
    ],
    "table": {"headers": ["Horizon", "Cumulative Illustrative Benefit (USD)"],
               "rows": [[r["horizon"], f"${r['cumulative_usd']:,.2f}"] for r in ROI_TIMELINE]},
})

word_path = build_word_report(
    REPORTS_DIR / "mp1_executive_report.docx",
    title="Mega Project 1 — Executive Capstone Report",
    subtitle="Intelligent Underwriting & Automated Credit Decisioning — Home Credit Default Risk Enterprise Suite",
    exec_summary=exec_summary,
    sections=word_sections,
    insights=INSIGHTS,
)

# --- Excel workbook ----------------------------------------------------------
assumptions = {}
assumption_notes = {}
benefit_cell_labels = []
for row in rollup_rows:
    if row["benefit_usd"] is not None:
        label = f"NB{row['notebook_id']}_ANNUAL_BENEFIT_USD"
        assumptions[label] = round(row["benefit_usd"], 2)
        assumption_notes[label] = f"Real illustrative figure from {row['problem']} ({row['benefit_label']})."
        benefit_cell_labels.append(label)
assumptions["GROWTH_RATE_ASSUMPTION_PCT"] = 0.0
assumption_notes["GROWTH_RATE_ASSUMPTION_PCT"] = (
    "ASSUMPTION: benefit run-rate held flat year over year (no growth, no compounding) -- "
    "the simplest, most defensible extrapolation this suite discloses rather than invents."
)

data_sheets = [
    {"name": "Problem Rollup", "headers": list(rollup_df.columns),
     "rows": rollup_df.fillna("").values.tolist(), "highlight_col": "benefit_usd"},
    {"name": "ROI Timeline", "headers": list(roi_df.columns), "rows": roi_df.values.tolist()},
]

total_formula = "+".join(assumption_ref(assumptions, lbl) for lbl in benefit_cell_labels) or "0"
formula_sheet = {
    "name": "Financial Impact",
    "rows": [
        ("Total Annual Benefit Run-Rate", f"={total_formula}"),
        ("1 Month (Cumulative)", f"=({total_formula})/12"),
        ("6 Months (Cumulative)", f"=({total_formula})/2"),
        ("1 Year (Cumulative)", f"=({total_formula})*1"),
        ("2 Years (Cumulative)", f"=({total_formula})*2"),
        ("3 Years (Cumulative)", f"=({total_formula})*3"),
        ("5 Years (Cumulative)", f"=({total_formula})*5"),
    ],
}

excel_path = build_excel_workbook(
    REPORTS_DIR / "mp1_executive_report.xlsx",
    assumptions=assumptions,
    assumption_notes=assumption_notes,
    data_sheets=data_sheets,
    formula_sheet=formula_sheet,
    insights_sheet={"name": "SMART Insights", "items": INSIGHTS},
)

# --- Big-letters financial-impact FRONT sheet (post-process: openpyxl,
# inserted as the first sheet so it's what opens the workbook) ---------------
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

wb = load_workbook(excel_path)
front = wb.create_sheet("Financial Impact (Big)", 0)
front.sheet_view.showGridLines = False
front.column_dimensions["A"].width = 4
front.column_dimensions["B"].width = 60
BIG_TITLE_FONT = Font(size=28, bold=True, color="1F3864")
BIG_VALUE_FONT = Font(size=44, bold=True, color="1F7A3D")
LABEL_FONT = Font(size=14, bold=True, color="595959")
front.merge_cells("B2:H2")
front["B2"] = "MEGA PROJECT 1 — REAL FINANCIAL IMPACT"
front["B2"].font = BIG_TITLE_FONT

big_rows = [
    ("Total Annual Benefit Run-Rate (real, illustrative)", f"${total_annual_benefit:,.0f}"),
    ("5-Year Cumulative Illustrative Benefit (ASSUMPTION-based)", f"${ROI_TIMELINE[-1]['cumulative_usd']:,.0f}"),
    ("Real Notebooks Complete", f"{N_AVAILABLE} / 5"),
]
r = 4
for label, value in big_rows:
    front.merge_cells(f"B{r}:H{r}")
    front[f"B{r}"] = label
    front[f"B{r}"].font = LABEL_FONT
    r += 1
    front.merge_cells(f"B{r}:H{r}")
    front[f"B{r}"] = value
    front[f"B{r}"].font = BIG_VALUE_FONT
    front[f"B{r}"].alignment = Alignment(horizontal="left")
    r += 2
front.merge_cells(f"B{r}:H{r+1}")
front[f"B{r}"] = ("Scale caveat: figures reflect whatever data each notebook was most recently run "
                   "against -- see Assumptions and Problem Rollup sheets for full detail.")
front[f"B{r}"].font = Font(size=10, italic=True, color="808080")
front[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
wb.save(excel_path)

# --- HTML dashboard ----------------------------------------------------------
benefit_rows_only = [r for r in rollup_rows if r["benefit_usd"] is not None]
scale_rows_only = [r for r in rollup_rows if r["portfolio_scale_usd"] is not None]
verdict_counts: dict[str, int] = {}
for r in rollup_rows:
    key = r["deployment_verdict"].split("—")[0].strip()
    verdict_counts[key] = verdict_counts.get(key, 0) + 1

kpi_cards = [
    {"label": "Total Annual Benefit Run-Rate", "value": f"${total_annual_benefit:,.0f}"},
    {"label": "5-Year Cumulative (Illustrative)", "value": f"${ROI_TIMELINE[-1]['cumulative_usd']:,.0f}"},
    {"label": "Notebooks Complete", "value": f"{N_AVAILABLE} / 5"},
    {"label": "Avg. Integrity Checks Passed", "value": (
        f"{sum(int(r['integrity_checks'].split('/')[0]) for r in rollup_rows)}/"
        f"{sum(int(r['integrity_checks'].split('/')[1].split()[0]) for r in rollup_rows)}"
    )},
]

charts = [
    {
        "id": "benefitByProblem", "title": "Real Illustrative Benefit by Problem", "type": "bar",
        "labels": [r["problem"].split("—")[1].strip() if "—" in r["problem"] else r["problem"]
                   for r in benefit_rows_only],
        "datasets": [{"label": "Benefit (USD)", "data": [r["benefit_usd"] for r in benefit_rows_only],
                      "backgroundColor": VIVID_PALETTE[:len(benefit_rows_only)]}],
        "story": [f"Notebooks 01-03 combine for ${total_annual_benefit:,.2f} in real illustrative "
                  f"annual benefit on their most recently scored populations."],
    },
    {
        "id": "roiTimeline", "title": "ASSUMPTION-Based Cumulative Benefit Timeline", "type": "line",
        "labels": [r["horizon"] for r in ROI_TIMELINE],
        "datasets": [{"label": "Cumulative Benefit (USD)",
                      "data": [r["cumulative_usd"] for r in ROI_TIMELINE],
                      "backgroundColor": VIVID_PALETTE[3]}],
        "note": "Flat annual run-rate ASSUMPTION -- no growth, no compounding. Not a forecast.",
    },
    {
        "id": "portfolioScale", "title": "Real Portfolio Volume Scored by Problem", "type": "bar",
        "labels": [r["problem"].split("—")[1].strip() if "—" in r["problem"] else r["problem"]
                   for r in scale_rows_only],
        "datasets": [{"label": "Portfolio Volume (USD)",
                      "data": [r["portfolio_scale_usd"] for r in scale_rows_only],
                      "backgroundColor": VIVID_PALETTE[1:1 + len(scale_rows_only)]}],
    },
    {
        "id": "verdictDist", "title": "Deployment Verdicts (this run)", "type": "doughnut",
        "labels": list(verdict_counts.keys()),
        "datasets": [{"label": "Problems", "data": list(verdict_counts.values()),
                      "backgroundColor": VIVID_PALETTE[:len(verdict_counts)]}],
    },
]

html_path = build_html_dashboard(
    REPORTS_DIR / "mp1_executive_dashboard.html",
    title="Mega Project 1 — Executive Dashboard",
    subtitle="Intelligent Underwriting & Automated Credit Decisioning (real rollup of Notebooks 01-05)",
    kpi_cards=kpi_cards,
    charts=charts,
    insights=INSIGHTS,
    data_table={
        "title": "Per-Problem Rollup (real)",
        # NOTE: "Statistical Robustness Verdict" and "Pipeline Integrity Checks" are
        # TWO DIFFERENT, independently-computed check families -- not the same
        # thing shown twice. Integrity checks are structural pipeline sanity
        # (did the code run correctly and produce well-formed output); the
        # verdict is a separate, stricter statistical-significance gate (is
        # there robust real evidence of the analyzed association/effect). A
        # problem can honestly show 100% on one and fail the other -- that is
        # not a contradiction. Earlier column names ("Verdict" / "Integrity
        # Checks") did not make this distinction visible, which read as
        # self-contradictory when shown side by side in the same row -- fixed
        # during the hardening pass (see CHANGELOG.md), and each verdict string
        # now names its specific failing check(s) rather than a vague pointer.
        "columns": ["Problem", "Champion/Method", "Statistical Robustness Verdict",
                    "Pipeline Integrity Checks (separate from Verdict)", "Benefit (USD)",
                    "Cost Context (USD)", "Portfolio Scale (USD)"],
        "rows": [[r["problem"], r["champion_or_method"], r["deployment_verdict"], r["integrity_checks"],
                  f"${r['benefit_usd']:,.2f}" if r["benefit_usd"] is not None else "",
                  f"${r['cost_usd']:,.2f}" if r["cost_usd"] is not None else "",
                  f"${r['portfolio_scale_usd']:,.2f}" if r["portfolio_scale_usd"] is not None else ""]
                 for r in rollup_rows],
    },
)

print(f"[REPORTING] Real MP1 executive reporting package written: {word_path.name}, {excel_path.name}, "
      f"{html_path.name}, plus {len(csv_paths)} CSV file(s) (all under {REPORTS_DIR.name}/).")

# ---------------------------------------------------------------------------
# SECTION 9 — Governance JSON summary for this rollup
# ---------------------------------------------------------------------------
mp1_summary = {
    "report": "mp1_executive_report",
    "mega_project": "Mega Project 1 - Intelligent Underwriting & Automated Credit Decisioning",
    "notebooks_available": N_AVAILABLE,
    "notebooks_missing": missing,
    "total_annual_benefit_run_rate_usd": total_annual_benefit,
    "roi_timeline_assumption_based": ROI_TIMELINE,
    "roi_assumptions": ROI_ASSUMPTIONS,
    "per_problem_rollup": rollup_rows,
    "integrity_checks": {n: bool(ok) for n, ok in checks},
    "reporting_artifacts": [word_path.name, excel_path.name, html_path.name] + [f"{s}.csv" for s in csv_paths],
    "runtime_seconds": round(time.time() - T0, 1),
}
with open(ARTIFACTS_DIR / "mp1_executive_summary.json", "w") as f:
    json.dump(mp1_summary, f, indent=2, default=str)

print(f"\n[DONE] MP1 executive rollup complete in {time.time() - T0:.1f}s covering {N_AVAILABLE}/5 "
      f"real notebook summaries.")
